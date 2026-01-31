import pandas as pd
import numpy as np
from datetime import datetime
import io
import re
from django.http import HttpResponse

# ---------------------------
# CONSTANTS
# ---------------------------
REQUIRED_COLUMNS = [
    "GSTIN/UIN", "Supplier", "Invoice", "Date",
    "Gross Amt", "Taxable", "IGST", "SGST", "CGST", "Cess", "Type"
]

NUMERIC_COLUMNS = ["Gross Amt", "Taxable", "IGST", "SGST", "CGST", "Cess"]

class GSTR2BManualReconciliationService:
    """
    Service for manual GSTR-2B vs Books reconciliation using two uploaded files.
    Ported from the 'reconciliation' app on 'main' branch.
    """

    @staticmethod
    def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
        df.columns = df.columns.astype(str).str.strip()
        if not df.columns.is_unique:
            counts = {}
            new_columns = []
            for col in df.columns:
                if col in counts:
                    counts[col] += 1
                    new_columns.append(f"{col}_{counts[col]}") 
                else:
                    counts[col] = 0
                    new_columns.append(col)
            df.columns = new_columns
        return df

    @staticmethod
    def preprocess_data(df: pd.DataFrame) -> pd.DataFrame:
        for col in NUMERIC_COLUMNS:
            if col not in df.columns:
                df[col] = 0
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        if "Invoice" in df.columns:
            df["Invoice"] = (
                df["Invoice"]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True) 
                .replace(["nan", "None", "NaN"], "")
            )
            df["Invoice_Clean"] = df["Invoice"].str.strip().str.upper()

        if "GSTIN/UIN" in df.columns:
            df["GSTIN/UIN"] = df["GSTIN/UIN"].astype(str).replace(["nan", "None"], "")
            df["GSTIN_Clean"] = df["GSTIN/UIN"].str.strip().str.upper()

        df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")

        if "Type" not in df.columns:
            df["Type"] = "B2B"
        
        df["Type"] = df["Type"].astype(str).str.strip().str.upper()
        cdnr_pattern = r"(CDNR|CREDIT|CR\.|DEBIT|DR\.|NOTE)"
        df.loc[df["Type"].str.contains(cdnr_pattern, regex=True, na=False), "Type"] = "CDNR"
        df.loc[df["Type"] != "CDNR", "Type"] = "B2B"

        return df

    @staticmethod
    def get_target_periods(selected_fy, period_type, selected_period_val):
        # selected_fy: "2024-2025"
        try:
            start_year = int(selected_fy.split("-")[0])
        except:
            start_year = 2024
            
        period_label = f"{selected_fy} - {selected_period_val}"
        
        month_map = {
            "April": 4, "May": 5, "June": 6, "July": 7, "August": 8, "September": 9,
            "October": 10, "November": 11, "December": 12, "January": 1, "February": 2, "March": 3
        }
        
        q_map = {
            "Q1 (Apr-Jun)": [4, 5, 6],
            "Q2 (Jul-Sep)": [7, 8, 9],
            "Q3 (Oct-Dec)": [10, 11, 12],
            "Q4 (Jan-Mar)": [1, 2, 3]
        }

        target_dates = []
        if period_type == "Monthly":
            m = month_map.get(selected_period_val, 4)
            y = start_year if m >= 4 else start_year + 1
            target_dates = [(y, m)]
        elif period_type == "Quarterly":
            ms = q_map.get(selected_period_val, [4,5,6])
            target_dates = [(start_year if m >= 4 else start_year + 1, m) for m in ms]
        else: # Yearly
            target_dates = [(start_year, m) for m in range(4, 13)] + [(start_year + 1, m) for m in range(1, 4)]
            period_label = f"{selected_fy} - Entire Year"

        return target_dates, period_label

    def run_reconciliation(self, df_2b, df_books, target_dates, tolerance=1):
        def in_period(dt, periods):
            if pd.isnull(dt): return False
            return (dt.year, dt.month) in periods

        # Filter Period
        mask_2b_in = df_2b["Date"].apply(lambda x: in_period(x, target_dates))
        df_2b_period = df_2b[mask_2b_in].copy()
        
        mask_books_in = df_books["Date"].apply(lambda x: in_period(x, target_dates))
        df_books_period = df_books[mask_books_in].copy()
        df_out_of_period = df_books[~mask_books_in].copy()

        results_matched = []
        results_mismatch_probable = []
        results_invoice_mismatch = []
        results_only_2b = []
        results_only_books = []

        # 1. Exact Match (GSTIN + Invoice)
        merged = pd.merge(
            df_2b_period, df_books_period, 
            on=["GSTIN_Clean", "Invoice_Clean"], 
            how="outer", suffixes=("_2b", "_books")
        )

        leftover_2b = merged[merged["Taxable_books"].isna()].copy()
        leftover_books = merged[merged["Taxable_2b"].isna()].copy()
        matched_df = merged[merged["Taxable_2b"].notna() & merged["Taxable_books"].notna()].copy()

        def safe_str(val):
            if pd.isna(val) or val is None: return ""
            return str(val)

        def coalesce_row(row, keys):
            for k in keys:
                if k in row and not pd.isna(row[k]) and row[k] != "":
                    return row[k]
            return ""

        for _, row in matched_df.iterrows():
            is_val_match = (
                abs(row.get("Taxable_2b", 0) - row.get("Taxable_books", 0)) <= tolerance and
                abs(row.get("IGST_2b", 0) - row.get("IGST_books", 0)) <= tolerance
            )
            
            base_data = {
                "GSTIN": safe_str(coalesce_row(row, ["GSTIN_Clean", "GSTIN/UIN_2b", "GSTIN/UIN_books"])),
                "Supplier": safe_str(coalesce_row(row, ["Supplier_2b", "Supplier_books"])),
                "Invoice_2B": safe_str(coalesce_row(row, ["Invoice_2b"])),
                "Invoice_Books": safe_str(coalesce_row(row, ["Invoice_books"])),
                "Date_2B": row.get("Date_2b"),
                "Date_Books": row.get("Date_books"),
                "Taxable_2B": row.get("Taxable_2b", 0),
                "Taxable_Books": row.get("Taxable_books", 0),
                "IGST_2B": row.get("IGST_2b", 0),
                "IGST_Books": row.get("IGST_books", 0),
                "CGST_2B": row.get("CGST_2b", 0),
                "CGST_Books": row.get("CGST_books", 0),
                "SGST_2B": row.get("SGST_2b", 0),
                "SGST_Books": row.get("SGST_books", 0),
                "Cess_2B": row.get("Cess_2b", 0),
                "Cess_Books": row.get("Cess_books", 0),
                "Gross_2B": row.get("Gross Amt_2b", 0),
                "Gross_Books": row.get("Gross Amt_books", 0),
                "Gross_Diff": round(abs(row.get("Gross Amt_2b", 0) - row.get("Gross Amt_books", 0)), 2),
                "Type": row.get("Type_2b", "B2B")
            }

            if is_val_match:
                results_matched.append(base_data)
            else:
                results_mismatch_probable.append(base_data)

        # 2. Fuzzy Match on leftover
        cols_2b = {col: col.replace("_2b", "") for col in leftover_2b.columns if "_2b" in col}
        cols_2b.update({"GSTIN_Clean": "GSTIN_Clean", "Invoice_Clean": "Invoice_Original_2B"})
        
        # Clean up column list for candidate DFs
        filtered_cols_2b = [c for c in leftover_2b.columns if c in cols_2b]
        df_2b_candidate = leftover_2b[filtered_cols_2b].rename(columns=cols_2b)

        cols_books = {col: col.replace("_books", "") for col in leftover_books.columns if "_books" in col}
        cols_books.update({"GSTIN_Clean": "GSTIN_Clean", "Invoice_Clean": "Invoice_Original_Books"})
        
        filtered_cols_books = [c for c in leftover_books.columns if c in cols_books]
        df_books_candidate = leftover_books[filtered_cols_books].rename(columns=cols_books)

        unmatched_2b_indices = set(df_2b_candidate.index)
        unmatched_books_indices = set(df_books_candidate.index)

        def values_match(v1, v2, tol):
            return abs(float(v1 or 0) - float(v2 or 0)) <= tol

        for idx_2b, row_2b in df_2b_candidate.iterrows():
            possible_books = df_books_candidate[df_books_candidate["GSTIN_Clean"] == row_2b["GSTIN_Clean"]]
            for idx_books, row_books in possible_books.iterrows():
                if idx_books not in unmatched_books_indices: continue

                if (values_match(row_2b.get("Taxable"), row_books.get("Taxable"), tolerance) and
                    values_match(row_2b.get("IGST"), row_books.get("IGST"), tolerance)):
                    
                    gross_m = values_match(row_2b.get("Gross Amt"), row_books.get("Gross Amt"), tolerance)
                    
                    match_data = {
                        "GSTIN": safe_str(row_2b.get("GSTIN_Clean")),
                        "Supplier": safe_str(row_2b.get("Supplier", "")),
                        "Invoice_2B": safe_str(row_2b.get("Invoice_Clean", "")),
                        "Invoice_Books": safe_str(row_books.get("Invoice_Clean", "")),
                        "Date_2B": row_2b.get("Date"),
                        "Date_Books": row_books.get("Date"),
                        "Taxable_2B": row_2b.get("Taxable", 0),
                        "Taxable_Books": row_books.get("Taxable", 0),
                        "IGST_2B": row_2b.get("IGST", 0),
                        "IGST_Books": row_books.get("IGST", 0),
                        "CGST_2B": row_2b.get("CGST", 0),
                        "CGST_Books": row_books.get("CGST", 0),
                        "SGST_2B": row_2b.get("SGST", 0),
                        "SGST_Books": row_books.get("SGST", 0),
                        "Cess_2B": row_2b.get("Cess", 0),
                        "Cess_Books": row_books.get("Cess", 0),
                        "Gross_2B": row_2b.get("Gross Amt", 0),
                        "Gross_Books": row_books.get("Gross Amt", 0),
                        "Gross_Diff": round(abs(row_2b.get("Gross Amt", 0) - row_books.get("Gross Amt", 0)), 2),
                        "Type": row_2b.get("Type", "B2B")
                    }

                    if gross_m: results_invoice_mismatch.append(match_data)
                    else: results_mismatch_probable.append(match_data)

                    unmatched_2b_indices.discard(idx_2b)
                    unmatched_books_indices.discard(idx_books)
                    break

        # orphans
        for idx in unmatched_2b_indices:
            row = df_2b_candidate.loc[idx]
            results_only_2b.append({
                "GSTIN": safe_str(row.get("GSTIN_Clean")),
                "Supplier": safe_str(row.get("Supplier")),
                "Invoice_2B": safe_str(row.get("Invoice_Clean")),
                "Invoice_Books": "", "Date_2B": row.get("Date"), "Date_Books": "",
                "Taxable_2B": row.get("Taxable", 0), "Taxable_Books": 0,
                "IGST_2B": row.get("IGST", 0), "IGST_Books": 0,
                "CGST_2B": row.get("CGST", 0), "CGST_Books": 0,
                "SGST_2B": row.get("SGST", 0), "SGST_Books": 0,
                "Cess_2B": row.get("Cess", 0), "Cess_Books": 0,
                "Gross_2B": row.get("Gross Amt", 0), "Gross_Books": 0,
                "Gross_Diff": 0, "Type": row.get("Type", "B2B")
            })

        for idx in unmatched_books_indices:
            row = df_books_candidate.loc[idx]
            results_only_books.append({
                "GSTIN": safe_str(row.get("GSTIN_Clean")),
                "Supplier": safe_str(row.get("Supplier")),
                "Invoice_2B": "", "Invoice_Books": safe_str(row.get("Invoice_Clean")),
                "Date_2B": "", "Date_Books": row.get("Date"),
                "Taxable_2B": 0, "Taxable_Books": row.get("Taxable", 0),
                "IGST_2B": 0, "IGST_Books": row.get("IGST", 0),
                "CGST_2B": 0, "CGST_Books": row.get("CGST", 0),
                "SGST_2B": 0, "SGST_Books": row.get("SGST", 0),
                "Cess_2B": 0, "Cess_Books": row.get("Cess", 0),
                "Gross_2B": 0, "Gross_Books": row.get("Gross Amt", 0),
                "Gross_Diff": 0, "Type": row.get("Type", "B2B")
            })

        return {
            "matched": pd.DataFrame(results_matched),
            "mismatch_probable": pd.DataFrame(results_mismatch_probable),
            "invoice_mismatch": pd.DataFrame(results_invoice_mismatch),
            "only_2b": pd.DataFrame(results_only_2b),
            "only_books": pd.DataFrame(results_only_books),
            "out_of_period": df_out_of_period
        }

    def generate_advanced_excel(self, results_dict, period_label):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        output = io.BytesIO()
        wb = Workbook()
        
        # --- STYLES ---
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_font = Font(bold=True, size=16, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        subheader_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        label_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        data_header_fill = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")

        # --- SUMMARY SHEET ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.merge_cells('A1:E2')
        cell = ws_summary['A1']
        cell.value = "GSTR-2B vs BOOKS RECONCILIATION SUMMARY"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        ws_summary['A4'] = "Period:"
        ws_summary['A4'].fill = label_fill
        ws_summary['A4'].border = thin_border
        ws_summary['B4'] = period_label
        ws_summary['B4'].fill = label_fill
        ws_summary['B4'].border = thin_border
        
        # Metrics
        ws_summary.merge_cells('A6:C6')
        ws_summary['A6'] = "KEY METRICS"
        ws_summary['A6'].fill = subheader_fill
        ws_summary['A6'].font = Font(bold=True, size=12, color="1E3A8A")
        ws_summary['A6'].border = thin_border
        
        color_map = {
            "matched": "DCFCE7", "mismatch_probable": "FFEDD5", "invoice_mismatch": "FEF9C3",
            "only_2b": "FEE2E2", "only_books": "F3E8FF", "out_of_period": "F1F5F9"
        }
        
        metrics_row = 7
        for k, color in color_map.items():
            df = results_dict.get(k)
            cnt = len(df) if df is not None else 0
            ws_summary.cell(row=metrics_row, column=1, value=k.replace("_", " ").title())
            ws_summary.cell(row=metrics_row, column=1).fill = label_fill
            ws_summary.cell(row=metrics_row, column=1).border = thin_border
            ws_summary.cell(row=metrics_row, column=2, value=cnt)
            ws_summary.cell(row=metrics_row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws_summary.cell(row=metrics_row, column=2).font = Font(bold=True)
            ws_summary.cell(row=metrics_row, column=2).border = thin_border
            metrics_row += 1

        # --- DATA SHEETS ---
        sheet_map = {
            "matched": ("Matched", "DCFCE7"),
            "mismatch_probable": ("Probable Mismatch", "FFEDD5"),
            "invoice_mismatch": ("Invoice Mismatch", "FEF9C3"),
            "only_2b": ("Only In 2B", "FEE2E2"),
            "only_books": ("Only In Books", "F3E8FF"),
            "out_of_period": ("Out Of Period", "F1F5F9")
        }

        for key, (sheet_title, color) in sheet_map.items():
            df = results_dict.get(key)
            if df is not None and not df.empty:
                ws = wb.create_sheet(title=sheet_title[:31])
                
                # Title row
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
                title_cell = ws.cell(row=1, column=1, value=sheet_title.upper())
                title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                title_cell.font = Font(bold=True)
                title_cell.alignment = Alignment(horizontal='center')
                title_cell.border = thin_border
                
                # Header row
                for c_idx, col in enumerate(df.columns, 1):
                    cell = ws.cell(row=2, column=c_idx, value=col)
                    cell.fill = data_header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    ws.column_dimensions[cell.column_letter].width = 15

                # Data rows
                for r_idx, row in enumerate(df.values, 3):
                    for c_idx, val in enumerate(row, 1):
                        col_name = df.columns[c_idx-1].lower()
                        cell = ws.cell(row=r_idx, column=c_idx)
                        
                        if "date" in col_name and pd.notnull(val) and val != "":
                            try:
                                cell.value = pd.to_datetime(val)
                                cell.number_format = 'DD-MM-YYYY'
                            except:
                                cell.value = str(val)
                        elif isinstance(val, (int, float)) and any(x in col_name for x in ["taxable", "igst", "cgst", "sgst", "gross", "cess"]):
                            cell.value = val
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = str(val) if val is not None else ""
                        
                        cell.border = thin_border

        wb.save(output)
        output.seek(0)
        return output
