# gst_reports/services/gstr1_excel_service.py

import pandas as pd
import io
import uuid
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gst_reports.services.gst_data_service import GSTDataService

UNWANTED_PREFIXES = (
    "status", "error", "timestamp", "transaction", "data.b2b.inv.itms.num",
    "data.b2b.inv.flag", "data.b2b.inv.updby", "data.b2b.inv.cflag", "data.cdnr.nt.cflag",
    "data.b2b.inv.chksum", "data.b2ba.inv.chksum", "data.b2cs.chksum", "data.b2csa.chksum",
    "data.cdnr.chksum", "data.cdnra.nt.chksum", "data.hsn.chksum", "data.exp.inv.chksum",
    "data.doc_issue.chksum", "data.b2b.cfs", "data.b2ba.cfs", "data.cdnra.cfs", "data.cdnr.cfs",
    "data.b2cs.flag", "data.b2csa.flag", "data.cdnr.nt.flag", "data.cdnra.nt.flag",
    "data.exp.inv.flag", "data.hsn.flag", "data.doc_issue.flag",
)

VALUE_MAPPING = {
    "Invoice Type": {"R": "Regular"},
    "Reverse Charge": {"Y": "Yes", "N": "No"},
}

ENDPOINTS = {
    "summary": "", "b2b": "b2b", "b2ba": "b2ba", "b2cl": "b2cl", "b2cla": "b2cla",
    "b2cs": "b2cs", "b2csa": "b2csa", "cdnr": "cdnr", "cdnra": "cdnra", "cdnur": "cdnur",
    "cdnura": "cdnura", "exp": "exp", "nil": "nil", "hsn": "hsn", "docs": "doc-issue",
    "at": "at", "ata": "ata"
}

OUTPUT_COLUMNS = [
    "Return Period", "Filing Status", "Original Invoice Number", "Original Invoice Date",
    "Invoice Number", "Invoice Date", "Invoice Value", "Place of Supply", "Reverse Charge",
    "E-Commerce GSTIN", "Invoice Type", "Applicable % of Tax Rate", "GSTIN/UIN of Recipient",
    "Receiver Name", "Rate", "Taxable Value", "Tax Amount", "IGST Amount", "CGST Amount",
    "SGST Amount", "CESS Amount", "IBN", "Generation Date", "Source Type",
    "Nature of Document", "Sr. No. From", "Sr. No. To", "Total Number", "Cancelled", "Net issued",
    "HSN", "Description", "UQC", "Total Quantity", "Total Value", "Section name", "Number of documents", "Total Amount"
]

COLUMN_MAPPING = {
    "ctin": "GSTIN/UIN of Recipient", "cname": "Receiver Name", "oinum": "Original Invoice Number",
    "oidt": "Original Invoice Date", "inum": "Invoice Number", "idt": "Invoice Date", "val": "Invoice Value",
    "pos": "Place of Supply", "rchrg": "Reverse Charge", "etin": "E-Commerce GSTIN", "inv_typ": "Invoice Type",
    "rt": "Rate", "txval": "Taxable Value", "iamt": "IGST Amount", "camt": "CGST Amount", "samt": "SGST Amount",
    "csamt": "CESS Amount", "nt_num": "Invoice Number", "nt_dt": "Invoice Date", "ont_num": "Original Invoice Number",
    "ont_dt": "Original Invoice Date", "sply_ty": "Supply Type", "exp_typ": "Export Type", "doc_desc": "Nature of Document", 
    "totnum": "Total Number", "cancel": "Cancelled", "net_issue": "Net issued", "hsn_sc": "HSN",
    "desc": "Description", "uqc": "UQC", "qty": "Total Quantity", "sec_nm": "Section name",
    "ttl_doc": "Number of documents", "ttl_tax": "Taxable Value", "ttl_igst": "IGST Amount",
    "ttl_cgst": "CGST Amount", "ttl_sgst": "SGST Amount", "ttl_cess": "CESS Amount", "ttl_val": "Total Amount"
}


class GSTR1ExcelService:
    @staticmethod
    def get_periods(download_type, fy=None, quarter=None, year=None, month=None):
        if download_type == "fy":
            y = int(fy.split("-")[0])
            now = datetime.now()
            months = []
            for m in range(4, 13):
                if y < now.year or (y == now.year and m <= now.month): months.append((y, m))
            for m in range(1, 4):
                if y + 1 < now.year or (y + 1 == now.year and m <= now.month): months.append((y + 1, m))
            return months, fy
        elif download_type == "quarterly":
            y = int(fy.split("-")[0])
            q = int(quarter)
            months = [(y, 4), (y, 5), (y, 6)] if q == 1 else \
                     [(y, 7), (y, 8), (y, 9)] if q == 2 else \
                     [(y, 10), (y, 11), (y, 12)] if q == 3 else \
                     [(y + 1, 1), (y + 1, 2), (y + 1, 3)]
            return months, f"{fy}_Q{quarter}"
        else:
            m = int(month)
            y = int(year)
            # If FY is provided, use it as the source of truth for the calendar year
            if fy:
                start_year = int(fy.split("-")[0])
                y = start_year if m >= 4 else start_year + 1
                
            return [(y, m)], f"{m:02d}{y}"

    @staticmethod
    def fetch_data(user, gstin, access_token, endpoint, year, month, force_refresh=False):
        section = endpoint if endpoint else "summary"
        if section == "summary":
            return GSTDataService.get_gstr1_summary(user, gstin, year, month, access_token, force_refresh)
        return GSTDataService.get_gstr1_section(user, gstin, section, year, month, access_token, force_refresh)

    @staticmethod
    def flatten_json(data, parent="", rows=None):
        if rows is None: rows = [{}]
        if isinstance(data, dict):
            for k, v in data.items(): rows = GSTR1ExcelService.flatten_json(v, f"{parent}.{k}" if parent else k, rows)
        elif isinstance(data, list):
            new_rows = []
            for item in data:
                for r in rows: new_rows.extend(GSTR1ExcelService.flatten_json(item, parent, [r.copy()]))
            rows = new_rows
        else:
            for r in rows: r[parent] = data
        return rows

    @staticmethod
    def clean_dataframe(df, sheet_name=""):
        if df.empty: return df
        df = df[[c for c in df.columns if not any(c.startswith(p) for p in UNWANTED_PREFIXES)]]
        renamed = {}
        for col in df.columns:
            for k, v in COLUMN_MAPPING.items():
                if k in col.split('.'):
                    renamed[col] = v
                    break
        df = df.rename(columns=renamed)
        if sheet_name == "hsn" and "Invoice Value" in df.columns: df = df.rename(columns={"Invoice Value": "Total Value"})
        df = df.loc[:, ~df.columns.duplicated()]
        for col, mapping in VALUE_MAPPING.items():
            if col in df.columns: df[col] = df[col].map(mapping).fillna(df[col])
        tax_cols = ["IGST Amount", "CGST Amount", "SGST Amount", "CESS Amount"]
        existing = [c for c in tax_cols if c in df.columns]
        if existing:
            for c in existing: df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
            df["Tax Amount"] = df[existing].sum(axis=1)
        present = [c for c in OUTPUT_COLUMNS if c in df.columns]
        return df[present + [c for c in df.columns if c not in present]].reset_index(drop=True)

    def generate(self, user, gstin, access_token, download_type, fy=None, quarter=None, year=None, month=None, force_refresh=False, username=None):
        months_list, period_label = self.get_periods(download_type, fy, quarter, year, month)
        sheets = {k: [] for k in ENDPOINTS}
        tasks = [(s, e, y, m) for y, m in months_list for s, e in ENDPOINTS.items()]
        
        errors = []
        results = {}
        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_task = {
                executor.submit(self.fetch_data, user, gstin, access_token, t[1], t[2], t[3], force_refresh): t for t in tasks
            }
            for future in as_completed(future_to_task):
                task = future_to_task[future]
                try:
                    results[task] = future.result()
                except Exception as e: 
                    errors.append(str(e))
        
        # Process results in original task order (preserving chronological sequence)
        for task in tasks:
            sheet, _, yr, mn = task
            data = results.get(task)
            if data:
                rows = self.flatten_json(data)
                for r in rows:
                    r.update({"Month": mn, "Return Period": f"{mn:02d}{yr}", "Filing Status": "FILED", "Source Type": "Portal"})
                sheets[sheet].extend(rows)
        
        if errors: raise Exception("Download failed: " + " | ".join(list(set(errors))[:3]))
        
        output = io.BytesIO()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        user_suffix = f"_{username}" if username else ""
        filename = f"GSTR1_{gstin}{user_suffix}_{period_label}_{timestamp}.xlsx"
        
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_count = 0
            for sheet, rows in sheets.items():
                if not rows: continue
                df = self.clean_dataframe(pd.DataFrame(rows), sheet_name=sheet)
                df.to_excel(writer, sheet_name=sheet, index=False)
                sheet_count += 1
                worksheet = writer.sheets[sheet]
                for col_num, value in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = blue_fill if "Original" in str(value) else orange_fill
                    cell.font = header_font
                    cell.border = thin_border
                    worksheet.column_dimensions[get_column_letter(col_num)].width = 20
            if sheet_count == 0: pd.DataFrame(columns=OUTPUT_COLUMNS).to_excel(writer, sheet_name="No Data", index=False)
        
        output.seek(0)
        return output, filename
