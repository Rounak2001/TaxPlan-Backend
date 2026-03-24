# gst_reports/services/gstr3b_excel_service.py

import io
import calendar
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gst_reports.services.gst_data_service import GSTDataService

class GSTR3BExcelService:
    @staticmethod
    def generate(user, gstin, year, month, taxpayer_token, force_refresh=False, username=None):
        data = GSTDataService.get_gstr3b_filed(
            user=user, gstin=gstin, year=year, month=month, taxpayer_token=taxpayer_token, force_refresh=force_refresh
        )
        if not data:
            month_name = calendar.month_name[month]
            raise Exception(f"No GSTR-3B data found for {month_name} {year}. The return may not be filed yet.")
        
        gstr3b = data
        if "data" in gstr3b and isinstance(gstr3b["data"], dict):
            gstr3b = gstr3b["data"]
            
        ret_period = gstr3b.get("ret_period", f"{month:02d}{year}")
        month_name = calendar.month_name[month]

        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-3B Summary"

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11, color="1F4E78")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        money_format = '₹#,##0.00'

        ws.merge_cells('A1:F1')
        ws['A1'] = f"GSTR-3B Details - {month_name} {year}"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center')
        user_info = f" | User: {username}" if username else ""
        ws['A2'] = f"GSTIN: {gstin}{user_info} | Return Period: {ret_period}"
        ws['A2'].font = Font(bold=True, size=11)

        row = 4
        # TABLE 3.1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, "TABLE 3.1 - OUTWARD SUPPLIES").fill = section_fill
        ws.cell(row, 1).font = section_font
        row += 1
        headers = ["Particulars", "Taxable Value", "IGST", "CGST", "SGST", "CESS"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row += 1
        sup_details = gstr3b.get("sup_details", {})
        outward_rows = [
            ("3.1(a) Outward taxable supplies", "osup_det"),
            ("3.1(b) Outward taxable supplies (zero rated)", "osup_zero"),
            ("3.1(c) Other outward supplies (Nil rated, exempted)", "osup_nil_exmp"),
            ("3.1(d) Inward supplies liable to reverse charge", "isup_rev"),
            ("3.1(e) Non-GST outward supplies", "osup_nongst"),
        ]
        for label, key in outward_rows:
            sec = sup_details.get(key, {})
            ws.cell(row, 1, label).border = border
            for c, k in enumerate(["txval", "iamt", "camt", "samt", "csamt"], 2):
                ws.cell(row, c, float(sec.get(k, 0) or 0)).number_format = money_format
                ws.cell(row, c).border = border
            row += 1
        
        row += 1
        # TABLE 4
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, "TABLE 4 - ELIGIBLE ITC").fill = section_fill
        ws.cell(row, 1).font = section_font
        row += 1
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row, col, h); cell.fill = header_fill; cell.font = header_font; cell.border = border
        row += 1
        itc_elg = gstr3b.get("itc_elg", {})
        itc_type_labels = {"IMPG": "4(A)(1) Import of goods", "IMPS": "4(A)(2) Import of services", "ISRC": "4(A)(3) Inward supplies liable to RCM", "ISD": "4(A)(4) Inward supplies from ISD", "OTH": "4(A)(5) All other ITC"}
        for itc_item in itc_elg.get("itc_avl", []):
            label = itc_type_labels.get(itc_item.get("ty", ""), f"4(A) {itc_item.get('ty', '')}")
            ws.cell(row, 1, label).border = border
            ws.cell(row, 2, "-").border = border
            for c, k in enumerate(["iamt", "camt", "samt", "csamt"], 3):
                ws.cell(row, c, float(itc_item.get(k, 0) or 0)).number_format = money_format
                ws.cell(row, c).border = border
            row += 1
        
        itc_net = itc_elg.get("itc_net", {})
        ws.cell(row, 1, "4(C) Net ITC Available").font = Font(bold=True)
        ws.cell(row, 1).border = border
        ws.cell(row, 2, "-").border = border
        for c, k in enumerate(["iamt", "camt", "samt", "csamt"], 3):
            cell = ws.cell(row, c, float(itc_net.get(k, 0) or 0))
            cell.number_format = money_format; cell.border = border; cell.font = Font(bold=True)
        row += 2

        # TABLE 6
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, "TABLE 6 - TAX PAYMENT").fill = section_fill; ws.cell(row, 1).font = section_font; row += 1
        pay_headers = ["Description", "IGST", "CGST", "SGST", "CESS", "Interest"]
        for col, h in enumerate(pay_headers, 1):
            cell = ws.cell(row, col, h); cell.fill = header_fill; cell.font = header_font; cell.border = border
        row += 1
        for item in gstr3b.get("tx_pmt", {}).get("net_tax_pay", []):
            ws.cell(row, 1, item.get("tran_desc", "")).border = border
            keys = [("igst", "tx"), ("cgst", "tx"), ("sgst", "tx"), ("cess", "tx"), ("igst", "intr")]
            for c, (k1, k2) in enumerate(keys, 2):
                val = item.get(k1, {}).get(k2, 0) if isinstance(item.get(k1), dict) else 0
                ws.cell(row, c, float(val or 0)).number_format = money_format
                ws.cell(row, c).border = border
            row += 1
        
        ws.column_dimensions['A'].width = 45
        for col in ['B', 'C', 'D', 'E', 'F']: ws.column_dimensions[col].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        user_suffix = f"_{username}" if username else ""
        filename = f"GSTR3B_Details_{gstin}{user_suffix}_{month_name}_{year}.xlsx"
        return output, filename

    @staticmethod
    def generate_multi(user, gstin, periods, taxpayer_token, force_refresh=False, username=None):
        """
        Generate a GSTR-3B Excel workbook for multiple periods (Option A).
        Months as Columns (Horizontal), Particulars as Rows.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-3B Combined Summary"

        # Fetch data for all periods in parallel
        results = {}
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_period = {
                executor.submit(
                    GSTDataService.get_gstr3b_filed, 
                    user, gstin, yr, mn, taxpayer_token, force_refresh
                ): (yr, mn) for yr, mn in periods
            }
            # results must be sorted by period
            for future in as_completed(future_to_period):
                period = future_to_period[future]
                try:
                    data = future.result()
                    if data:
                        if "data" in data and isinstance(data["data"], dict):
                            data = data["data"]
                        results[period] = data
                except Exception as e:
                    print(f"Error fetching data for {period}: {e}")

        available_periods = [p for p in periods if p in results]
        if not available_periods:
            raise Exception("No GSTR-3B data found for any of the selected periods.")

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11, color="1F4E78")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        money_format = '#,##0.00'

        total_cols = 1 + (len(available_periods) * 5)

        # Title and Header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws['A1'] = "GSTR-3B Combined Comparative Report (Option A)"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78"); ws['A1'].alignment = Alignment(horizontal='center')
        
        user_info = f" | User: {username}" if username else ""
        period_range = f"{calendar.month_name[available_periods[0][1]]} {available_periods[0][0]} to {calendar.month_name[available_periods[-1][1]]} {available_periods[-1][0]}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        ws['A2'] = f"GSTIN: {gstin}{user_info} | Period: {period_range}"
        ws['A2'].font = Font(bold=True, size=11); ws['A2'].alignment = Alignment(horizontal='center')

        # Header Row 4: Months
        row_h4 = 4
        ws.cell(row_h4, 1, "Particulars / Sections").fill = header_fill; ws.cell(row_h4, 1).font = header_font; ws.cell(row_h4, 1).border = border
        for i, (yr, mn) in enumerate(available_periods):
            start_col = 2 + (i * 5)
            month_label = f"{calendar.month_name[mn][:3]}-{yr}"
            ws.merge_cells(start_row=row_h4, start_column=start_col, end_row=row_h4, end_column=start_col + 4)
            cell = ws.cell(row_h4, start_col, month_label)
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center'); cell.border = border
        
        # Header Row 5: Fields (Taxable Value, IGST, etc.)
        row_h5 = 5
        ws.cell(row_h5, 1, "").border = border
        headers = ["Taxable Value", "IGST", "CGST", "SGST", "CESS"]
        for i in range(len(available_periods)):
            start_col = 2 + (i * 5)
            for j, h in enumerate(headers):
                cell = ws.cell(row_h5, start_col + j, h)
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center'); cell.border = border
        
        row_idx = 6

        def write_comparative_section(cur_row, title, label_keys, data_path_keys):
            # Section Header
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=total_cols)
            ws.cell(cur_row, 1, title).fill = section_fill; ws.cell(cur_row, 1).font = section_font
            cur_row += 1
            
            for label, key in label_keys:
                ws.cell(cur_row, 1, label).border = border; ws.cell(cur_row, 1).font = Font(bold=False)
                for i, period in enumerate(available_periods):
                    start_col = 2 + (i * 5)
                    gstr3b = results.get(period, {})
                    
                    sec_data = gstr3b
                    for p in data_path_keys:
                        sec_data = sec_data.get(p, {}) if isinstance(sec_data, dict) else {}
                    
                    row_data = sec_data.get(key, {}) if isinstance(sec_data, dict) else {}
                    for j, f_key in enumerate(["txval", "iamt", "camt", "samt", "csamt"]):
                        val = float(row_data.get(f_key, 0) or 0)
                        ws.cell(cur_row, start_col + j, val).number_format = money_format; ws.cell(cur_row, start_col + j).border = border
                cur_row += 1
            return cur_row

        # Table 3.1
        row_idx = write_comparative_section(row_idx, "TABLE 3.1 - OUTWARD SUPPLIES", [
            ("3.1(a) Outward taxable supplies", "osup_det"),
            ("3.1(b) Outward taxable supplies (zero)", "osup_zero"),
            ("3.1(c) Other outward supplies (Nil/Exemp)", "osup_nil_exmp"),
            ("3.1(d) Inward supplies (RCM)", "isup_rev"),
            ("3.1(e) Non-GST outward supplies", "osup_nongst"),
        ], ["sup_details"])
        
        row_idx += 1
        
        # Table 4 - Eligible ITC
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        ws.cell(row_idx, 1, "TABLE 4 - ELIGIBLE ITC").fill = section_fill; ws.cell(row_idx, 1).font = section_font
        row_idx += 1
        
        itc_labels = [
            ("4(A)(1) Import of goods", "IMPG"),
            ("4(A)(2) Import of services", "IMPS"),
            ("4(A)(3) Inward supplies liable to RCM", "ISRC"),
            ("4(A)(4) Inward supplies from ISD", "ISD"),
            ("4(A)(5) All other ITC", "OTH"),
        ]
        
        for label, ty_code in itc_labels:
            ws.cell(row_idx, 1, label).border = border
            for i, period in enumerate(available_periods):
                start_col = 2 + (i * 5)
                itc_avl = results.get(period, {}).get("itc_elg", {}).get("itc_avl", [])
                target = next((x for x in itc_avl if x.get("ty") == ty_code), {})
                # Taxable is N/A for ITC
                ws.cell(row_idx, start_col, "-").border = border; ws.cell(row_idx, start_col).alignment = Alignment(horizontal='center')
                for j, f_key in enumerate(["iamt", "camt", "samt", "csamt"], 1):
                    val = float(target.get(f_key, 0) or 0)
                    ws.cell(row_idx, start_col + j, val).number_format = money_format; ws.cell(row_idx, start_col + j).border = border
            row_idx += 1
            
        # Net ITC
        ws.cell(row_idx, 1, "4(C) Net ITC Available").font = Font(bold=True); ws.cell(row_idx, 1).border = border
        for i, period in enumerate(available_periods):
            start_col = 2 + (i * 5)
            itc_net = results.get(period, {}).get("itc_elg", {}).get("itc_net", {})
            ws.cell(row_idx, start_col, "-").border = border; ws.cell(row_idx, start_col).alignment = Alignment(horizontal='center')
            for j, f_key in enumerate(["iamt", "camt", "samt", "csamt"], 1):
                val = float(itc_net.get(f_key, 0) or 0)
                ws.cell(row_idx, start_col + j, val).number_format = money_format; ws.cell(row_idx, start_col + j).border = border; ws.cell(row_idx, start_col + j).font = Font(bold=True)
        row_idx += 2
        
        # Table 6 - Cash Payment
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        ws.cell(row_idx, 1, "TABLE 6 - TAX PAYMENT IN CASH").fill = section_fill; ws.cell(row_idx, 1).font = section_font
        row_idx += 1
        
        for label, key in [("Integrated Tax", "igst"), ("Central Tax", "cgst"), ("State/UT Tax", "sgst"), ("Cess", "cess")]:
            ws.cell(row_idx, 1, label).border = border
            for i, period in enumerate(available_periods):
                start_col = 2 + (i * 5)
                # In Table 6, Taxable is N/A. Sub-fields are IGST, CGST... but for each row only one matters usually.
                # However, to maintain the grid, we use the same columns.
                tx_pmt = results.get(period, {}).get("tx_pmt", {}).get("net_tax_pay", [])
                val = 0
                for item in tx_pmt:
                    desc = item.get("tran_desc", "").lower()
                    if key.lower() in desc or (key == "igst" and "integrated" in desc):
                        val = float(item.get(key, {}).get("tx", 0) or 0)
                        break
                ws.cell(row_idx, start_col, "-").border = border; ws.cell(row_idx, start_col).alignment = Alignment(horizontal='center')
                # Put the value in its respective column for the month? 
                # Actually, let's just use the columns IGST, CGST, SGST, CESS relative to start_col
                field_indices = {"igst": 1, "cgst": 2, "sgst": 3, "cess": 4}
                for j in range(1, 5):
                    v = val if j == field_indices[key] else 0
                    ws.cell(row_idx, start_col + j, v).number_format = money_format; ws.cell(row_idx, start_col + j).border = border
            row_idx += 1

        # Column widths
        ws.column_dimensions['A'].width = 45
        for i in range(2, total_cols + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        user_suffix = f"_{username}" if username else ""
        filename = f"GSTR3B_Combined_{gstin}{user_suffix}.xlsx"
        return output, filename
        
    # Note: Accidental overwrites are fixed. Restored generate/generate_multi correctly.
