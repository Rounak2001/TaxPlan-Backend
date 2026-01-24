import calendar
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReconciliationExcelService:
    @staticmethod
    def generate_1vs3b_excel(results, username, gstin, fy_year):
        wb = Workbook()
        
        # --- Styles ---
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        month_font = Font(bold=True, color="FFFFFF", size=11)
        diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ========== SHEET 0: Summary View ==========
        ws_sum = wb.active
        ws_sum.title = "Consolidated Summary"
        
        ws_sum.merge_cells('A1:G1')
        ws_sum['A1'] = "GST Health Check - Consolidated Summary"
        ws_sum['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws_sum['A1'].alignment = Alignment(horizontal='center')
        
        headers = ["Month", "Liability (R1)", "Payment (3B)", "Diff (L)", "Credit (2B)", "Payment (3B)", "Diff (C)"]
        for col, h in enumerate(headers, 1):
            cell = ws_sum.cell(3, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row = 4
        import calendar
        for data in results:
            m_name = calendar.month_name[data['month']] + " " + str(data['year'])
            
            # Liability
            l_r1 = float(data.get('ig1', 0) or 0) + float(data.get('cg1', 0) or 0) + float(data.get('sg1', 0) or 0)
            l_3b = float(data.get('ig3', 0) or 0) + float(data.get('cg3', 0) or 0) + float(data.get('sg3', 0) or 0)
            l_diff = l_r1 - l_3b
            
            # Credit
            c_2b = float(data.get('itc_2b_igst', 0) or 0) + float(data.get('itc_2b_cgst', 0) or 0) + float(data.get('itc_2b_sgst', 0) or 0)
            c_3b = float(data.get('itc_adj_igst', 0) or 0) + float(data.get('itc_adj_cgst', 0) or 0) + float(data.get('itc_adj_sgst', 0) or 0)
            c_diff = c_2b - c_3b
            
            # Write Row
            vals = [m_name, l_r1, l_3b, l_diff, c_2b, c_3b, c_diff]
            for col, val in enumerate(vals, 1):
                c = ws_sum.cell(row, col, val)
                c.border = border
                if col > 1:
                    c.number_format = '#,##0.00'
                    if col in (4, 7): # Diff columns
                        if abs(val) > 1:
                            c.fill = diff_fill
                        else:
                            c.fill = match_fill
            row += 1
            
        ws_sum.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws_sum.column_dimensions[col].width = 15

        def create_reco_sheet(ws, title, particulars, subtitle):
            """Helper to create a reconciliation sheet"""
            total_cols = max(len(results) * 4 + 1, 5)
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws['A2'] = f"Username: {username} | GSTIN: {gstin} | FY: {fy_year}-{int(fy_year) + 1}"
            ws['A2'].font = Font(bold=True, size=11)
            
            ws['A4'] = "Particular"
            ws['A4'].fill = header_fill
            ws['A4'].font = header_font
            ws['A4'].border = border
            
            # Month Headers
            col = 2
            for data in results:
                month_name = calendar.month_abbr[data['month']] + " " + str(data['year'])
                ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+2)
                cell = ws.cell(4, col, month_name)
                cell.fill = month_fill
                cell.font = month_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                ws.cell(5, col, subtitle[0]).font = Font(bold=True, size=9)
                ws.cell(5, col+1, subtitle[1]).font = Font(bold=True, size=9)
                ws.cell(5, col+2, "Diff").font = Font(bold=True, size=9)
                col += 4
            
            # Data Rows
            row = 6
            for particular, key_auto, key_filed in particulars:
                ws.cell(row, 1, particular).border = border
                col = 2
                for data in results:
                    auto_val = float(data.get(key_auto, 0) or 0)
                    filed_val = float(data.get(key_filed, 0) or 0)
                    diff = auto_val - filed_val
                    
                    c1 = ws.cell(row, col, round(auto_val, 2))
                    c1.number_format = '#,##0.00'
                    c1.border = border
                    
                    c2 = ws.cell(row, col+1, round(filed_val, 2))
                    c2.number_format = '#,##0.00'
                    c2.border = border
                    
                    c3 = ws.cell(row, col+2, round(diff, 2))
                    c3.number_format = '#,##0.00'
                    c3.border = border
                    
                    if abs(diff) > 1:
                        c3.fill = diff_fill
                        c3.font = Font(bold=True, color="9C0006")
                    else:
                        c3.fill = match_fill
                        c3.font = Font(color="006100")
                    
                    col += 4
                row += 1
            
            ws.column_dimensions['A'].width = 30
            for i in range(2, col):
                ws.column_dimensions[get_column_letter(i)].width = 18
            ws.freeze_panes = 'B6'

        # ========== SHEET 1: Sales (GSTR-1 vs GSTR-3B) ==========
        ws_sales = wb.create_sheet("Sales (R1 vs 3B)")
        
        sales_particulars = [
            ('3.1.a Taxable Value', 'tx1', 'tx3'),
            ('3.1.a IGST', 'ig1', 'ig3'),
            ('3.1.a CGST', 'cg1', 'cg3'),
            ('3.1.a SGST', 'sg1', 'sg3'),
            ('3.1.b Export Taxable', 'exp_tx1', 'exp_tx3'),
            ('3.1.b Export IGST', 'exp_ig1', 'exp_ig3'),
            ('3.1.c Nil/Exempt', 'nil_tx1', 'nil_tx3'),
            ('3.1.e Non-GST', 'ng1', 'ng3'),
        ]
        create_reco_sheet(ws_sales, "GSTR-1 vs GSTR-3B Reconciliation (Sales)", 
                          sales_particulars, ("GSTR-1", "GSTR-3B"))
        
        # ========== SHEET 2: Purchases (GSTR-2B vs GSTR-3B ITC) ==========
        ws_purchases = wb.create_sheet("Purchases (2B vs 3B)")
        
        itc_particulars = [
            ('ITC - IGST', 'itc_2b_igst', 'itc_adj_igst'),
            ('ITC - CGST', 'itc_2b_cgst', 'itc_adj_cgst'),
            ('ITC - SGST', 'itc_2b_sgst', 'itc_adj_sgst'),
            ('ITC - CESS', 'itc_2b_cess', 'itc_adj_cess'),
        ]
        create_reco_sheet(ws_purchases, "GSTR-2B vs GSTR-3B ITC Reconciliation (RCM Adjusted)", 
                          itc_particulars, ("GSTR-2B", "GSTR-3B (Adj)"))
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"GSTR_Reconciliation_{gstin}_{fy_year}.xlsx"
        return output, filename

    @staticmethod
    def generate_books_reco_excel(results_data, username, gstin, year, title):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Handle if results_data is the full dict or just the summary list
        if isinstance(results_data, dict):
            results = results_data.get('summary', [])
            all_results = results_data
        else:
            results = results_data
            all_results = {}

        # Header Info
        ws.merge_cells('A1:Z1')
        ws['A1'] = f"{title} | Username: {username} | GSTIN: {gstin} | Year/FY: {year}"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = center_align

        # Extract particulars from first month block
        particulars = []
        if results and len(results) > 0 and 'rows' in results[0]:
            particulars = [r['particular'] for r in results[0]['rows']]
        
        if not particulars:
            # Fallback if structure is missing
            ws.cell(row=3, column=1, value="No data available")
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output, f"Reconciliation_{gstin}.xlsx"

        # Start writing headers
        ws.cell(row=3, column=1, value="Particular").font = Font(bold=True)
        ws.cell(row=3, column=1).fill = header_fill
        ws.cell(row=3, column=1).font = header_font
        ws.cell(row=3, column=1).border = border
        
        col_idx = 2
        for m_block in results:
            month_name = m_block['month']
            ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+2)
            cell = ws.cell(row=3, column=col_idx, value=month_name)
            cell.font = header_font
            cell.fill = month_fill
            cell.alignment = center_align
            cell.border = border
            
            # Sub-headers (Books, Portal, Diff)
            portal_label = "Portal"
            if "1vsbooks" in title.lower() or "GSTR1" in title: portal_label = "GSTR-1"
            elif "3bvsbooks" in title.lower() or "GSTR3B" in title: portal_label = "GSTR-3B"

            for i, label in enumerate(["Books", portal_label, "Diff"]):
                c = ws.cell(row=4, column=col_idx + i, value=label)
                c.font = Font(bold=True, size=9)
                c.border = border
                c.alignment = center_align
            
            col_idx += 4
        
        # Write data rows
        row_idx = 5
        for p_name in particulars:
            ws.cell(row=row_idx, column=1, value=p_name).border = border
            
            col_idx = 2
            for m_block in results:
                # Find matching row in this month block
                row_data = next((r for r in m_block['rows'] if r['particular'] == p_name), None)
                if row_data:
                    v1 = float(row_data.get('v1', 0) or 0)
                    v2 = float(row_data.get('v2', 0) or 0)
                    diff = float(row_data.get('diff', 0) or 0)
                    
                    for i, val in enumerate([v1, v2, diff]):
                        c = ws.cell(row=row_idx, column=col_idx + i, value=round(val, 2))
                        c.number_format = '#,##0.00'
                        c.border = border
                        # Highlight diff if > 1
                        if i == 2 and abs(diff) > 1:
                            c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                col_idx += 4
            row_idx += 1
            
        ws.column_dimensions['A'].width = 35
        for i in range(2, col_idx):
            ws.column_dimensions[get_column_letter(i)].width = 15

        # --- Detailed Sheets (if available) ---
        sections = ["B2B", "B2CL", "B2CS", "EXP", "SEZ", "CDNR"]
        header_map = {
            "Taxable_BOOKS": "Books Taxable", "IGST_BOOKS": "Books IGST", "CGST_BOOKS": "Books CGST", "SGST_BOOKS": "Books SGST",
            "Taxable_PORTAL": "Portal Taxable", "IGST_PORTAL": "Portal IGST", "CGST_PORTAL": "Portal CGST", "SGST_PORTAL": "Portal SGST",
            "Taxable_DIFF": "Difference Taxable", "IGST_DIFF": "Difference IGST", "CGST_DIFF": "Difference CGST", "SGST_DIFF": "Difference SGST"
        }

        import pandas as pd
        for section in sections:
            records = all_results.get(section, [])
            if records:
                detail_ws = wb.create_sheet(title=f"Detailed_{section}")
                df = pd.DataFrame(records)
                
                # Ensure specific columns come first
                cols = list(df.columns)
                priority = ["Year", "Month", "Status"]
                ordered_cols = [c for c in priority if c in cols] + [c for c in cols if c not in priority]
                df = df[ordered_cols]
                
                # Rename columns for display
                display_cols = [header_map.get(c, c) for c in df.columns]
                
                # Header Style
                for c_idx, col_name in enumerate(display_cols, 1):
                    cell = detail_ws.cell(row=1, column=c_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                    
                # Data and Formatting
                for r_idx, row_values in enumerate(df.values, 2):
                    for c_idx, value in enumerate(row_values, 1):
                        col_name = df.columns[c_idx-1]
                        
                        # Standard cell writing
                        cell = detail_ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                        
                        # Type-specific formatting
                        raw_col = col_name.lower()
                        is_financial = any(x in raw_col for x in ["taxable", "igst", "cgst", "sgst", "diff"])
                        
                        if is_financial and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                        elif "year" in raw_col or "month" in raw_col:
                            cell.number_format = '0'
                        elif "pos" in raw_col:
                            cell.number_format = '@'

                        # Highlight mismatches
                        if (col_name == "Status" and value == "Mismatch") or \
                           ("_DIFF" in col_name and isinstance(value, (int, float)) and abs(value) > 1.0):
                            cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                        elif (col_name == "Status" and value == "Matched") or \
                             ("_DIFF" in col_name and isinstance(value, (int, float)) and abs(value) <= 1.0):
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                # Auto-adjust column widths
                for i, col in enumerate(display_cols, 1):
                    max_length = max(len(str(col)), 10) + 4
                    detail_ws.column_dimensions[get_column_letter(i)].width = max_length

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        clean_title = title.replace(" ", "_").replace("/", "_")
        filename = f"{clean_title}_{gstin}_{year}.xlsx"
        return output, filename
